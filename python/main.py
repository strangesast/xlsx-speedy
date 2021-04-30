import os
import warnings
from pathlib import Path
from openpyxl import load_workbook

from pprint import pprint

warnings.simplefilter("ignore")

HEADER = [
    ("CUSTOMER", "customer"),
    ("TOP ASSY\nDESCRIPTION", "top_assembly_description"),
    ("TOP ASSY \nP/N", "top_assembly_part_number"),
    ("SALES \nORDER #", "order_number"),
    ("QTY", "qty"),
    ("DUE \nDATE", "due_date"),
    (".", ""),
    ("PART\nDESCRIPTION", "part_description"),
    ("PART NUMBER", "part_number"),
    ("MATERIALS", "materials"),
    ("SAW", "status_saw"),
    ("LATHE", "status_lathe"),
    ("MILL", "status_mill"),
    ("WELD", "status_weld"),
    ("Outside\nService", "status_outside_service"),
    ("PACK / ASSY", "status_assembly"),
    ("LOCATION", "location"),
    ("NOTES", "notes"),
]


STATUS_MAP = {
    "Ä": 3,
    "X": 2,
    "/": 1,
    "..": 0,
    ".": 0,
}


def map_value(v, k: str):
    if k.startswith("status"):
        return (
            {"status": sv}
            if (type(v) is str and (sv := STATUS_MAP.get(v.upper())))
            else v
        )
    return v


EMPTY_THRESHOLD = 4


def fn(filepath, sheet_name):
    # fucking slow without read_only
    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb[sheet_name]

    it = enumerate(ws.iter_rows(min_row=0, max_col=18))
    header_row = next(r for j, r in it if j == 2)
    for col_num, (a, b) in enumerate(
        zip((c.value for c in header_row), (h[0] for h in HEADER))
    ):
        assert a == b, f"invalid header! {col_num=} {repr(a)} != {repr(b)}"

    header_keys = [h for _, h in HEADER]
    empty_cnt = 0
    for row_idx, row in it:
        if row[0]:
            empty_cnt = 0

            ob = {
                k: (map_value(cell.value, k), cell.value, (col_idx, row_idx))
                for col_idx, (cell, k) in enumerate(zip(row, header_keys))
                if k
            }
            yield ob

        else:
            empty_cnt += 1

            if empty_cnt >= EMPTY_THRESHOLD:
                break

    wb.close()


def test():
    FILEPATH = os.environ.get("FILEPATH") or Path.cwd() / "../schedule.xlsm"
    SHEET_NAME = "PRODUCTION SCHEDULE"

    arr = list(fn(FILEPATH, SHEET_NAME))


if __name__ == "__main__":
    import timeit

    print(timeit.timeit("test()", setup="from __main__ import test", number=1))
